# -*- coding: utf-8 -*-
"""
CSV에서 지역·음악/피아노 학원 조건으로 필터링 후 Excel로 저장합니다.

실행 방법 (Windows, Python 3.11):
  cd D:\\Git\\Git\\academy_csv_search
  pip install -r requirements.txt
  python main.py

실행 후 프롬프트에 검색할 지역명을 입력합니다. (예: 안산시, 강남구)
결과 xlsx는 D:\\Git\\dataset\\{지역}.xlsx 형태로 저장됩니다.
(여러 지역은 안산시_수원시.xlsx 처럼 stem만 '_'로 연결)

시작 시 open.neis.go.kr 학원교습소정보(File) API와 비교해, 더 새 기준일 CSV가 있으면
INPUT_CSV_PATH 와 같은 폴더에 {YYYYMMDD} data.csv 로 받아 사용합니다.
로컬이 이미 같거나 더 최신이면 다운로드하지 않습니다.
"""

from __future__ import annotations

import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional

import pandas as pd

# ---------------------------------------------------------------------------
# 경로·설정 상수 (환경에 맞게 변경 가능)
# ---------------------------------------------------------------------------
# 로컬 기준 데이터 CSV (동기화 시 같은 폴더에 {YYYYMMDD} data.csv 로 저장)
INPUT_CSV_PATH = Path(r"D:\Git\20260228 data.csv")

# 나이스 교육정보 개방 포털 — 학원교습소정보 파일 탭
# https://open.neis.go.kr/portal/data/service/selectServicePage.do?infId=OPEN19220231012134453534385&infSeq=3
NEIS_SYNC_ENABLED = True
NEIS_BASE_URL = "https://open.neis.go.kr"
NEIS_INF_ID = "OPEN19220231012134453534385"
NEIS_INF_SEQ = "3"
NEIS_SEARCH_FILE_DATA_URL = f"{NEIS_BASE_URL}/portal/data/file/searchFileData.do"
NEIS_DOWNLOAD_BLOB_URL = f"{NEIS_BASE_URL}/portal/data/file/downloadBlobFileData.do"
NEIS_HTTP_TIMEOUT_SEC = 300
NEIS_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

OUTPUT_DIR = Path(r"D:\Git\dataset")
# 저장 파일명: {지역}.xlsx (지역은 sanitize_filename_stem 적용)
OUTPUT_XLSX_EXTENSION = ".xlsx"
EXCEL_SHEET_NAME = "search_result"
# 파일명 stem이 비었거나 너무 길 때
OUTPUT_FILENAME_FALLBACK_STEM = "academy_search"
OUTPUT_FILENAME_MAX_STEM_LENGTH = 150

# 업종 판별 키워드 (하나라도 포함되면 대상)
MUSIC_PIANO_KEYWORDS: tuple[str, ...] = (
    "피아노학원",
    "피아노 교습소",
    "피아노교습소",
    "음악학원",
    "음악 교습소",
    "음악교습소",
    "피아노",
    "음악",
    "실용음악",
    "재즈피아노",
    "클래식피아노",
)

# 등록상태: True이면 개원만 포함 (컬럼이 있을 때만 적용)
FILTER_ACTIVE_STATUS_ONLY = True
# 허용할 등록상태명 (기본: 개원만)
ACTIVE_STATUS_VALUES: tuple[str, ...] = ("개원",)

# CSV 로딩
CSV_ENCODINGS: tuple[str, ...] = ("utf-8-sig", "utf-8", "cp949", "euc-kr")
DEFAULT_SEPARATOR = ","

# 논리 컬럼명 → CSV에 없을 때 탐색할 후보 이름 (정확 일치 우선 후 부분 매칭)
COLUMN_ALIAS_CANDIDATES: dict[str, tuple[str, ...]] = {
    "행정구역명": ("행정구역명", "행정구역", "지역명"),
    "도로명주소": ("도로명주소",),
    "도로명상세주소": ("도로명상세주소", "상세주소"),
    "학원교습소명": ("학원교습소명",),
    "학원명": ("학원명",),
    "분야명": ("분야명",),
    "교습계열명": ("교습계열명",),
    "교습과정목록명": ("교습과정목록명",),
    "교습과정명": ("교습과정명",),
    "등록상태명": ("등록상태명",),
    "도로명우편번호": ("도로명우편번호",),
}

# 필수로 매핑되어야 하는 논리 컬럼 (필터·출력에 모두 사용)
REQUIRED_LOGICAL_COLUMNS: tuple[str, ...] = (
    "행정구역명",
    "도로명주소",
    "도로명상세주소",
    "학원교습소명",
    "학원명",
    "분야명",
    "교습계열명",
    "교습과정목록명",
    "교습과정명",
    "도로명우편번호",
)

# 최종 Excel 컬럼 순서
OUTPUT_COLUMN_ORDER: tuple[str, ...] = (
    "행정구역명",
    "학원교습소명",
    "학원명",
    "분야명",
    "교습계열명",
    "교습과정목록명",
    "교습과정명",
    "도로명주소",
    "도로명상세주소",
    "도로명우편번호",
)

KEYWORD_FILTER_LOGICAL_COLUMNS: tuple[str, ...] = (
    "학원교습소명",
    "학원명",
    "분야명",
    "교습계열명",
    "교습과정목록명",
    "교습과정명",
)

REGION_SEARCH_LOGICAL_COLUMNS: tuple[str, ...] = (
    "행정구역명",
    "도로명주소",
    "도로명상세주소",
)

DEDUP_KEY_LOGICAL_COLUMNS: tuple[str, ...] = (
    "학원명",
    "도로명주소",
    "도로명상세주소",
)

# 로컬에 있는 기준일 데이터 파일명: "20260228 data.csv"
_LOCAL_DATA_CSV_NAME_RE = re.compile(r"^(\d{8})\s+data\.csv$", re.IGNORECASE)


def _http_post_form(url: str, fields: dict[str, str]) -> bytes:
    """application/x-www-form-urlencoded POST 본문을 받습니다."""
    body = urllib.parse.urlencode(fields).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "User-Agent": NEIS_USER_AGENT,
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Origin": NEIS_BASE_URL,
            "Referer": (
                f"{NEIS_BASE_URL}/portal/data/service/selectServicePage.do"
                f"?infId={NEIS_INF_ID}&infSeq={NEIS_INF_SEQ}"
            ),
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def _http_get(url: str) -> bytes:
    req = urllib.request.Request(
        url,
        headers={"User-Agent": NEIS_USER_AGENT},
    )
    with urllib.request.urlopen(req, timeout=NEIS_HTTP_TIMEOUT_SEC) as resp:
        return resp.read()


def parse_baseline_date_from_neis_viewname(view_file_nm: str) -> Optional[date]:
    """
    API의 viewFileNm(예: 학원교습소정보_2026년02월28일기준, ...2025년_10월31일기준)에서
    기준일을 파싱합니다.
    """
    s = normalize_text_value(view_file_nm)
    m = re.search(r"_(\d{4})년_?(\d{1,2})월_?(\d{1,2})일", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def date_to_compact(d: date) -> str:
    """date → YYYYMMDD 문자열."""
    return f"{d.year:04d}{d.month:02d}{d.day:02d}"


def fetch_neis_academy_file_catalog() -> list[dict[str, Any]]:
    """
    searchFileData.do 로 파일 목록(JSON)을 가져옵니다.
    응답 형식: {"data": [{infId, infSeq, fileSeq, viewFileNm, fileExt, viewCnt}, ...]}
    """
    raw = _http_post_form(
        NEIS_SEARCH_FILE_DATA_URL,
        {"infId": NEIS_INF_ID, "infSeq": NEIS_INF_SEQ},
    )
    payload = json.loads(raw.decode("utf-8"))
    rows = payload.get("data")
    if not isinstance(rows, list):
        return []
    return [r for r in rows if isinstance(r, dict)]


def pick_latest_neis_csv_entry(rows: list[dict[str, Any]]) -> Optional[dict[str, Any]]:
    """
    CSV 항목 중 기준일(viewFileNm)이 가장 늦은 항목을 고릅니다.
    날짜 파싱 실패 시 fileSeq 최댓값으로 대체합니다.
    """
    csv_rows = [
        r
        for r in rows
        if str(r.get("fileExt", "")).lower() == "csv"
    ]
    if not csv_rows:
        return None

    def sort_key(r: dict[str, Any]) -> tuple[int, int]:
        vn = str(r.get("viewFileNm", ""))
        bd = parse_baseline_date_from_neis_viewname(vn)
        compact = int(date_to_compact(bd)) if bd else 0
        fs = r.get("fileSeq")
        try:
            fsi = int(fs) if fs is not None else 0
        except (TypeError, ValueError):
            fsi = 0
        return (compact, fsi)

    return max(csv_rows, key=sort_key)


def scan_local_baseline_csv_files(data_dir: Path) -> dict[str, Path]:
    """
    data_dir 안의 'YYYYMMDD data.csv' 파일을 찾아 {YYYYMMDD: Path} 맵을 만듭니다.
    """
    out: dict[str, Path] = {}
    if not data_dir.is_dir():
        return out
    for p in data_dir.iterdir():
        if not p.is_file():
            continue
        m = _LOCAL_DATA_CSV_NAME_RE.match(p.name)
        if m:
            out[m.group(1)] = p
    return out


def local_max_baseline_compact(local_map: dict[str, Path]) -> tuple[Optional[str], Optional[Path]]:
    """로컬 맵에서 가장 큰 YYYYMMDD와 해당 경로를 반환합니다."""
    if not local_map:
        return None, None
    best = max(local_map.keys(), key=lambda s: int(s))
    return best, local_map[best]


def download_neis_csv_blob(file_seq: int, dest_path: Path) -> None:
    """
    downloadBlobFileData.do 에 GET으로 요청해 CSV 바이너리를 dest_path에 저장합니다.
    임시 파일에 쓴 뒤 교체합니다.
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    q = urllib.parse.urlencode(
        {
            "infId": NEIS_INF_ID,
            "infSeq": NEIS_INF_SEQ,
            "fileSeq": str(file_seq),
        }
    )
    url = f"{NEIS_DOWNLOAD_BLOB_URL}?{q}"
    data = _http_get(url)
    tmp = dest_path.with_suffix(dest_path.suffix + ".download_part")
    try:
        tmp.write_bytes(data)
        tmp.replace(dest_path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def sync_neis_academy_csv_source() -> tuple[Path, list[str]]:
    """
    나이스 포털의 최신 학원교습소 CSV와 로컬을 비교·동기화합니다.

    반환: (읽을 CSV 경로, 사용자 피드백 문장 목록)
    - 원격이 더 최신이면 data_dir 에 {YYYYMMDD} data.csv 로 저장 후 그 경로 사용
    - 동일/로컬이 더 최신이면 다운로드 생략 후 적절한 로컬 경로 사용
    """
    feedback: list[str] = []
    data_dir = INPUT_CSV_PATH.parent

    if not NEIS_SYNC_ENABLED:
        feedback.append("[안내] NEIS 자동 동기화가 꺼져 있습니다(NEIS_SYNC_ENABLED=False). 설정된 로컬 파일을 사용합니다.")
        return INPUT_CSV_PATH, feedback

    local_map = scan_local_baseline_csv_files(data_dir)
    local_compact, local_path = local_max_baseline_compact(local_map)

    # 설정 파일명에 날짜가 있으면 맵에 반영 (패턴과 다를 때 대비)
    m_cfg = re.search(r"(\d{8})", INPUT_CSV_PATH.name)
    if m_cfg and INPUT_CSV_PATH.is_file():
        local_map.setdefault(m_cfg.group(1), INPUT_CSV_PATH)
        local_compact, local_path = local_max_baseline_compact(local_map)

    try:
        catalog = fetch_neis_academy_file_catalog()
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError, OSError) as e:
        feedback.append(f"[경고] 나이스 파일 목록을 가져오지 못했습니다: {e}")
        if INPUT_CSV_PATH.is_file():
            feedback.append(f"[안내] 기존 로컬 파일을 사용합니다: {INPUT_CSV_PATH}")
            return INPUT_CSV_PATH, feedback
        if local_path and local_path.is_file():
            feedback.append(f"[안내] 기존 로컬 파일을 사용합니다: {local_path}")
            return local_path, feedback
        print("오류: 네트워크 실패이며 사용할 로컬 CSV가 없습니다.", file=sys.stderr)
        sys.exit(1)

    latest = pick_latest_neis_csv_entry(catalog)
    if not latest:
        feedback.append("[경고] 나이스에서 CSV 파일 항목을 찾지 못했습니다.")
        if INPUT_CSV_PATH.is_file():
            return INPUT_CSV_PATH, feedback
        if local_path and local_path.is_file():
            return local_path, feedback
        print("오류: 원격·로컬 모두 사용할 CSV가 없습니다.", file=sys.stderr)
        sys.exit(1)

    vn = str(latest.get("viewFileNm", ""))
    bd = parse_baseline_date_from_neis_viewname(vn)
    if bd is None:
        feedback.append(f"[경고] 원본 파일명에서 기준일을 파싱하지 못했습니다: {vn}")
        try:
            fs = int(latest["fileSeq"])
        except (KeyError, TypeError, ValueError):
            feedback.append("[경고] fileSeq 가 없어 동기화를 건너뜁니다.")
            return INPUT_CSV_PATH if INPUT_CSV_PATH.is_file() else (local_path or INPUT_CSV_PATH), feedback
        remote_compact = None
    else:
        remote_compact = date_to_compact(bd)
        fs = int(latest["fileSeq"])

    feedback.append(
        f"[나이스] 최신 공개 CSV(포털 기준): {vn} (fileSeq={fs})"
    )

    if remote_compact is None:
        feedback.append("[안내] 기준일 비교 없이 로컬 파일을 유지합니다.")
        return INPUT_CSV_PATH if INPUT_CSV_PATH.is_file() else (local_path or INPUT_CSV_PATH), feedback

    dest = data_dir / f"{remote_compact} data.csv"

    if local_compact is not None and int(local_compact) >= int(remote_compact):
        use: Optional[Path] = None
        if remote_compact in local_map:
            use = local_map[remote_compact]
        elif local_compact in local_map:
            use = local_map[local_compact]
        if use is None or not use.is_file():
            use = INPUT_CSV_PATH if INPUT_CSV_PATH.is_file() else local_path
        feedback.append(
            f"[안내] 로컬 데이터가 이미 최신입니다 (로컬 기준일 {local_compact}, "
            f"포털 기준일 {remote_compact}). 다운로드하지 않습니다."
        )
        feedback.append(f"[안내] 사용 파일: {use}")
        return use, feedback

    try:
        feedback.append(
            f"[안내] 포털에 더 새 데이터가 있습니다 (로컬 {local_compact or '없음'} → {remote_compact}). 다운로드합니다…"
        )
        download_neis_csv_blob(fs, dest)
        size_b = dest.stat().st_size
        size_mb = size_b / (1024 * 1024)
        feedback.append(
            f"[다운로드 완료] 나이스 포털 최신 CSV를 받아 저장했습니다: {dest}"
        )
        feedback.append(
            f"[다운로드 완료] 기준일 {remote_compact}, 용량 약 {size_mb:.2f} MB ({size_b:,} bytes)"
        )
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError) as e:
        feedback.append(f"[경고] 다운로드 실패: {e}")
        if INPUT_CSV_PATH.is_file():
            feedback.append(f"[안내] 기존 로컬 파일로 계속합니다: {INPUT_CSV_PATH}")
            return INPUT_CSV_PATH, feedback
        if local_path and local_path.is_file():
            feedback.append(f"[안내] 기존 로컬 파일로 계속합니다: {local_path}")
            return local_path, feedback
        print("오류: 다운로드에 실패했고 대체 로컬 파일이 없습니다.", file=sys.stderr)
        sys.exit(1)

    return dest, feedback


def ensure_output_dir() -> None:
    """
    출력 디렉터리가 없으면 생성합니다.
    부모 경로까지 포함해 mkdir(parents=True)로 처리합니다.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_text_value(value: Any) -> str:
    """
    단일 셀 값을 문자열로 정규화합니다.
    NaN/None/공백만 있는 문자열은 빈 문자열로 통일하고, 앞뒤 공백을 제거합니다.
    """
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none", "<na>"):
        return ""
    return s


def sanitize_filename_stem(name: str) -> str:
    """
    Windows에서 파일명에 쓸 수 없는 문자를 제거·치환하고 길이를 제한합니다.
    여러 지역을 '_'로 이은 문자열을 그대로 넘기면 됩니다.
    """
    invalid = '<>:"/\\|?*'
    s = normalize_text_value(name)
    for ch in invalid:
        s = s.replace(ch, "_")
    s = s.strip().rstrip(".")
    s = re.sub(r"_+", "_", s)
    if not s:
        return OUTPUT_FILENAME_FALLBACK_STEM
    if len(s) > OUTPUT_FILENAME_MAX_STEM_LENGTH:
        s = s[:OUTPUT_FILENAME_MAX_STEM_LENGTH].rstrip("._ ")
    if not s:
        return OUTPUT_FILENAME_FALLBACK_STEM
    return s


def build_output_xlsx_path(region_queries: list[str]) -> Path:
    """
    검색에 사용한 지역 문자열을 반영한 xlsx 경로를 만듭니다.
    쉼표로 여러 지역을 넣은 경우 stem은 '_'로 연결합니다.
    """
    if region_queries:
        stem_source = "_".join(region_queries)
    else:
        stem_source = OUTPUT_FILENAME_FALLBACK_STEM
    safe_stem = sanitize_filename_stem(stem_source)
    return OUTPUT_DIR / f"{safe_stem}{OUTPUT_XLSX_EXTENSION}"


def _read_csv_attempt(
    path: Path,
    encoding: str,
    sep: Optional[str],
) -> pd.DataFrame:
    """한 번의 read_csv 시도 (sep=None이면 pandas python engine이 구분자 추정)."""
    kwargs: dict[str, Any] = {
        "encoding": encoding,
        "dtype": str,
        "keep_default_na": True,
    }
    if sep is not None:
        kwargs["sep"] = sep
        kwargs["engine"] = "c"
    else:
        kwargs["sep"] = None
        kwargs["engine"] = "python"
        kwargs["on_bad_lines"] = "warn"
    return pd.read_csv(path, **kwargs)


def load_csv_with_fallback(csv_path: Path) -> pd.DataFrame:
    """
    CSV를 인코딩·구분자 후보 순으로 읽습니다.
    1) utf-8-sig, utf-8, cp949, euc-kr 각각에 대해 쉼표 구분 시도
    2) 같은 인코딩에서 열이 1개뿐이거나 예외가 나면 sep=None + python engine 자동 감지 시도
    파일이 없으면 메시지 출력 후 sys.exit(1) 합니다.
    """
    if not csv_path.is_file():
        print(f"오류: 입력 CSV 파일을 찾을 수 없습니다.\n경로: {csv_path}", file=sys.stderr)
        sys.exit(1)

    last_error: Optional[Exception] = None

    for enc in CSV_ENCODINGS:
        # 1차: 쉼표 구분 + C engine (대용량에 유리)
        try:
            df = _read_csv_attempt(csv_path, enc, DEFAULT_SEPARATOR)
            # 열이 1개뿐이면 구분자 오인 가능성이 있어 자동 감지로 넘김
            if df.shape[1] > 1:
                print(f"CSV 로드 성공: 인코딩={enc}, 구분자='{DEFAULT_SEPARATOR}' (C engine)")
                return df
        except Exception as e:
            last_error = e
        # 2차: python engine이 스니퍼로 구분자 추정
        try:
            df = _read_csv_attempt(csv_path, enc, None)
            if df.shape[1] >= 1:
                print(f"CSV 로드 성공: 인코딩={enc}, 구분자=자동 감지 (python engine)")
                return df
        except Exception as e:
            last_error = e

    print(
        "오류: 지원 인코딩·구분자 조합으로 CSV를 읽지 못했습니다.",
        file=sys.stderr,
    )
    if last_error:
        print(f"마지막 예외: {last_error}", file=sys.stderr)
    sys.exit(1)


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    모든 셀에 대해 normalize_text_value를 적용한 복사본을 반환합니다.
    원본은 변경하지 않습니다.
    """
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(normalize_text_value)
    return out


def _exact_match_column(
    available: Iterable[str],
    candidates: tuple[str, ...],
) -> Optional[str]:
    """후보 중 DataFrame에 정확히 있는 컬럼명을 반환."""
    avail_set = set(available)
    for c in candidates:
        if c in avail_set:
            return c
    return None


def _fuzzy_match_column(
    available: Iterable[str],
    candidates: tuple[str, ...],
) -> Optional[str]:
    """
    정확 일치가 없을 때, 사용 가능한 컬럼명에 후보 문자열이 부분 포함되는지 검사합니다.
    (예: '상세주소' in '도로명상세주소'는 후보가 컬럼에 포함될 때만 매칭)
    """
    cols = list(available)
    for cand in candidates:
        for col in cols:
            if cand in col or col in cand:
                return col
    return None


def map_target_columns(df_columns: Iterable[str]) -> dict[str, str]:
    """
    논리 컬럼명 → 실제 DataFrame 컬럼명 매핑을 구축합니다.
    정확 일치를 우선하고, 없으면 부분 문자열 유사 매칭을 사용합니다.
    (동일 실제 컬럼이 두 논리명에 매핑되지 않도록 후보 순서가 중요합니다.)
    """
    cols = list(df_columns)
    mapping: dict[str, str] = {}
    for logical, candidates in COLUMN_ALIAS_CANDIDATES.items():
        actual = _exact_match_column(cols, candidates)
        if actual is None:
            actual = _fuzzy_match_column(cols, candidates)
        if actual is not None:
            mapping[logical] = actual
    return mapping


def validate_required_columns(column_map: Mapping[str, str]) -> None:
    """
    필수 논리 컬럼이 모두 매핑되었는지 검사합니다.
    누락 시 메시지 출력 후 종료합니다.
    """
    missing = [name for name in REQUIRED_LOGICAL_COLUMNS if name not in column_map]
    if missing:
        print("오류: 다음 필수 컬럼을 CSV에서 찾을 수 없습니다:", file=sys.stderr)
        for m in missing:
            print(f"  - {m}", file=sys.stderr)
        sys.exit(1)


def _rename_to_logical(df: pd.DataFrame, column_map: Mapping[str, str]) -> pd.DataFrame:
    """실제 컬럼명을 논리명으로 통일한 DataFrame을 반환합니다."""
    inv: dict[str, str] = {actual: logical for logical, actual in column_map.items()}
    out = df.rename(columns=inv)
    return out


def filter_by_region(
    df: pd.DataFrame,
    region_queries: list[str],
    column_map: Mapping[str, str],
) -> pd.DataFrame:
    """
    region_queries 중 하나라도 행정구역명/도로명주소/도로명상세주소 컬럼 값에 포함되면 통과합니다.
    쿼리는 앞뒤 공백 제거 후 사용합니다.
    """
    if df.empty:
        return df
    queries = [normalize_text_value(q) for q in region_queries if normalize_text_value(q)]
    if not queries:
        return df.iloc[0:0].copy()

    logical_df = _rename_to_logical(df, column_map)
    mask = pd.Series(False, index=df.index)
    for logical in REGION_SEARCH_LOGICAL_COLUMNS:
        if logical not in logical_df.columns:
            continue
        col = logical_df[logical].astype(str)
        for q in queries:
            mask = mask | col.str.contains(q, regex=False, na=False)
    return df.loc[mask].copy()


def _row_matches_keywords(row: pd.Series, keywords: tuple[str, ...]) -> bool:
    """한 행의 지정 논리 컬럼 연결 문자열에 키워드가 하나라도 포함되는지."""
    parts: list[str] = []
    for logical in KEYWORD_FILTER_LOGICAL_COLUMNS:
        if logical in row.index:
            parts.append(str(row[logical]))
    text = "".join(parts)
    return any(k in text for k in keywords)


def filter_by_music_piano_keywords(
    df: pd.DataFrame,
    column_map: Mapping[str, str],
    keywords: tuple[str, ...] = MUSIC_PIANO_KEYWORDS,
) -> pd.DataFrame:
    """
    학원교습소명, 학원명, 분야명, 교습계열, 교습과정 관련 컬럼에
    음악/피아노 키워드가 하나라도 있으면 통과합니다.
    """
    if df.empty:
        return df
    logical_df = _rename_to_logical(df, column_map)
    mask = logical_df.apply(
        lambda r: _row_matches_keywords(r, keywords),
        axis=1,
    )
    return df.loc[mask].copy()


def filter_by_active_status(
    df: pd.DataFrame,
    column_map: Mapping[str, str],
    *,
    only_active: bool = FILTER_ACTIVE_STATUS_ONLY,
    active_values: tuple[str, ...] = ACTIVE_STATUS_VALUES,
) -> pd.DataFrame:
    """
    등록상태명 컬럼이 매핑되어 있고 only_active이 True이면
    active_values에 포함된 행만 남깁니다.
    컬럼이 없거나 only_active가 False이면 필터를 적용하지 않습니다.
    """
    if df.empty or not only_active:
        return df.copy()
    if "등록상태명" not in column_map:
        return df.copy()
    logical_df = _rename_to_logical(df, column_map)
    col = logical_df["등록상태명"].map(normalize_text_value)
    mask = col.isin(list(active_values))
    return df.loc[mask].copy()


def _drop_all_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """모든 셀이 빈 문자열인 행을 제거합니다."""
    if df.empty:
        return df
    nonempty = (df != "").any(axis=1)
    return df.loc[nonempty].copy()


def clean_result(
    df: pd.DataFrame,
    column_map: Mapping[str, str],
) -> pd.DataFrame:
    """
    논리 컬럼 기준으로 정규화 후 출력 컬럼만 선택하고,
    완전 빈 행 제거, 행 전체 중복 제거,
    (학원명+도로명주소+도로명상세주소) 기준 중복 제거를 수행합니다.
    """
    if df.empty:
        logical_empty = pd.DataFrame(columns=list(OUTPUT_COLUMN_ORDER))
        return logical_empty

    logical_df = _rename_to_logical(df, column_map)
    # 문자열 재정규화 (필터 단계 이후 값 보정)
    for col in logical_df.columns:
        logical_df[col] = logical_df[col].map(normalize_text_value)

    keep_cols = [c for c in OUTPUT_COLUMN_ORDER if c in logical_df.columns]
    out = logical_df[keep_cols].copy()
    out = _drop_all_empty_rows(out)
    out = out.drop_duplicates()
    # 동일 학원·동일 주소는 1건만 (다른 열 차이로 남은 중복 제거)
    dedup_keys = [k for k in DEDUP_KEY_LOGICAL_COLUMNS if k in out.columns]
    if dedup_keys:
        out = out.drop_duplicates(subset=dedup_keys, keep="first")
    return out


def save_to_excel(df: pd.DataFrame, path: Path, sheet_name: str = EXCEL_SHEET_NAME) -> None:
    """
    DataFrame을 xlsx로 저장하고 헤더 굵게, 첫 행 고정, 자동 필터, 열 너비를 조정합니다.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
        ws.freeze_panes = "A2"
        if df.shape[0] > 0:
            ws.auto_filter.ref = ws.dimensions
        elif ws.max_column:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}1"

        for idx, col in enumerate(df.columns, start=1):
            series = df[col].astype(str)
            max_len = max(
                [len(str(col))]
                + [len(s) for s in series.tolist()[:5000]],
                default=len(str(col)),
            )
            width = min(max(max_len + 2, 10), 60)
            ws.column_dimensions[get_column_letter(idx)].width = width


def _print_column_mapping(column_map: dict[str, str]) -> None:
    """논리 컬럼 → 실제 CSV 컬럼 매핑을 콘솔에 출력합니다."""
    print("컬럼 매핑 (논리명 → 실제 CSV 컬럼명):")
    for logical in sorted(column_map.keys()):
        print(f"  {logical} → {column_map[logical]}")


def _print_sample(df: pd.DataFrame, n: int = 5) -> None:
    """최종 결과 샘플을 콘솔에 출력합니다."""
    print(f"\n최종 샘플 (최대 {n}건):")
    if df.empty:
        print("  (데이터 없음)")
        return
    sample = df.head(n)
    with pd.option_context("display.max_columns", None, "display.width", 200, "display.max_colwidth", 40):
        print(sample.to_string(index=False))


def main() -> None:
    """입력 CSV 로드 → 필터 → 정제 → Excel 저장까지 오케스트레이션합니다."""
    print("=" * 60)
    print("학원 CSV 지역·음악/피아노 검색 도구")
    print("=" * 60)

    raw_region = input("검색할 지역명을 입력하세요 (예: 안산시, 강남구): ").strip()
    if not raw_region:
        print("오류: 지역명이 비어 있습니다.", file=sys.stderr)
        sys.exit(1)

    # 확장: 여러 지역은 쉼표로 구분해 OR 조건 검색 (예: "안산시, 수원시")
    region_queries = [
        normalize_text_value(x) for x in raw_region.split(",") if normalize_text_value(x)
    ]

    ensure_output_dir()

    source_csv, sync_feedback = sync_neis_academy_csv_source()
    print("\n--- 원본 데이터 동기화 ---")
    for line in sync_feedback:
        print(line)

    df_raw = load_csv_with_fallback(source_csv)
    print("\nCSV 로드 후 실제 컬럼 목록:")
    for i, c in enumerate(df_raw.columns.tolist(), 1):
        print(f"  {i:3}. {c}")

    column_map = map_target_columns(df_raw.columns)
    _print_column_mapping(column_map)
    validate_required_columns(column_map)

    n_total = len(df_raw)
    df_norm = normalize_dataframe(df_raw)

    df_region = filter_by_region(df_norm, region_queries, column_map)
    n_after_region = len(df_region)

    df_music = filter_by_music_piano_keywords(df_region, column_map)
    n_after_music = len(df_music)

    df_status = filter_by_active_status(df_music, column_map)
    n_after_status = len(df_status)

    df_final = clean_result(df_status, column_map)
    n_final = len(df_final)

    print("\n--- 검증 로그 ---")
    print(f"1. 전체 원본 행 수: {n_total}")
    print(f"2. 지역 필터 적용 후 행 수: {n_after_region}")
    print(f"3. 업종(음악/피아노) 필터 적용 후 행 수: {n_after_music}")
    print(f"4. 등록상태 필터 적용 후 행 수: {n_after_status}")
    print(f"5. 최종 정제 후 행 수: {n_final}")

    _print_sample(df_final, 5)

    output_xlsx_path = build_output_xlsx_path(region_queries)
    save_to_excel(df_final, output_xlsx_path, EXCEL_SHEET_NAME)

    print(f"\n저장 완료: {output_xlsx_path}")
    print(f"결과 건수: {n_final}")
    if n_final == 0:
        print("검색 결과 없음 (빈 시트 구조로 파일은 저장되었습니다.)")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n중단되었습니다.", file=sys.stderr)
        sys.exit(130)
